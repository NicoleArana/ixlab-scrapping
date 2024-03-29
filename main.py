from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from os.path import exists
import pickle
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
import pandas as pd 
from openpyxl.utils.dataframe import dataframe_to_rows
from time import sleep
import os

from env import USERNAME, PASSWORD

options = Options()

# options.add_argument("--headless")


driver = webdriver.Chrome(options=options)

driver.implicitly_wait(5)

def parse_time(s: str):
    if s == "-":
        return 0
    
    mins_replaced = s.replace(" min", ",")
    secs_replaced = mins_replaced.replace(" sec", "")
    parts = secs_replaced.split(",")

    if len(parts) == 1:
        return int(parts[0])
    if len(parts) == 2:
        if parts[1] == "":
            return int(parts[0]) * 60
        else:
            return int(parts[0]) * 60 + int(parts[1])
        
    


        

def login():
    if exists("zoho_cookies.pkl") and exists("ixlab_cookies.pkl"):
        driver.implicitly_wait(0)
        driver.get("https://www.zoho.com/es-xl/")
        cookies = pickle.load(open("zoho_cookies.pkl", "rb"))
        # We load the zoho credentials
        for c in cookies:
            if c.get("domain") == ".zoho.com":
                driver.add_cookie(c)
        driver.get("https://accounts.zoho.com/signin")
        for c in cookies:
            if c.get("domain") == "accounts.zoho.com":
                driver.add_cookie(c)
        driver.refresh()
        # we load the tec credentials
        driver.get("https://ixlab-mx.trainercentral.com/adminpage")
        cookies = pickle.load(open("ixlab_cookies.pkl", "rb"))
        for c in cookies:
            if "trainercentral.com" in c.get("domain"):
                driver.add_cookie(c)
        driver.implicitly_wait(15)
        driver.refresh()
    else:
        driver.get("https://accounts.zoho.com/signin?servicename=TrainerCentral&serviceurl=https%3A%2F%2Fixlab-mx.trainercentral.com%2Fadminpage")
        login_input = driver.find_element(By.ID, "login_id")
        next_btn = driver.find_element(By.ID, "nextbtn")
        password_input = driver.find_element(By.ID, "password")

        login_input.send_keys(USERNAME)
        next_btn.click()
        password_input.send_keys(PASSWORD)
        next_btn.click()

        # save zoho cookies
        driver.get("https://www.zoho.com/es-xl/")
        pickle.dump(driver.get_cookies(), open("zoho_cookies.pkl", "wb"))

        # save ixlab cookies
        driver.get("https://ixlab-mx.trainercentral.com/adminpage")
        pickle.dump(driver.get_cookies(), open("ixlab_cookies.pkl", "wb"))
 

def main():
    #Nos dirige al dashboard con todos los cursos
    driver.get("https://ixlab-mx.trainercentral.com/adminpage#/portal/727165297/courses")

    # encontramos cada curso y los guardamos en una lista de cursos
    courses = driver.find_elements(By.CSS_SELECTOR, ".item-wrapper")

    #creamos un empty dataframe 
     


    for i in range(len(courses)):
        #por cada curso en la lista le damos click 
        new_courses = driver.find_elements(By.CSS_SELECTOR, ".item-wrapper")
        current_course = new_courses[i]

        course_status = current_course.find_element(By.CSS_SELECTOR, "span[data-tctest=course_published_status]")
        print("Status: ", course_status.text)
        if (course_status.text == "Borrador"):
            continue

        # SKIPS FIRST ENTRIES FOR DEBUGGING
        # DELETE LATER
        if i < 5:
            continue

        current_course.click()

        # nos dirigimos a la pagina de plan de estudios en cada curso
        current_url = driver.current_url
        driver.get(current_url.replace("details", "addcurriculum"))   

        #buscamos los modulos dentro de cada curso y los guardamos en una lista
        # store these in a database
        modules_data = {
            "modulo": [],
            "sub_modulo": [],
            "tipo de archivo": [],
            "duracion": [],
        }


        # for course in course_modules:
        #     modules_data["modulo"].append(course.text)


        # driver.execute_script("arguments[0].scrollIntoView();", course_modules[-1])

        #buscamos los modulos dentro de cada curso
        course_title = driver.find_element(By.CSS_SELECTOR, ".breadcrumb-course-title")

        #Escribimos en el archivo el nombre del curso
        print(course_title.text)

        sleep(3)

        #por cada modulo en la lista de modulos escribimos su nombre en ese archivo

        #buscamos aquellos modulos que contengan un elemento video, le damos click y obtenemos la duración 
        # del video y la escribimos al archivo 
        
        # wait until
        #content-loader is-load
        
                
        sub_modules = driver.find_elements(By.CSS_SELECTOR, "tr.ember-view")
        submodules_len_modules = driver.find_elements(By.CSS_SELECTOR, "ul[class=list-unstyled]")    

        li_submodules_len_modules = []
        for s, time in enumerate(submodules_len_modules):
            if (s == len(submodules_len_modules) - 1):
                break 
            if (s % 2 == 0):
                li_submodules_len_modules.append(time.find_element(By.TAG_NAME, "li"))
                # try:
                #     li_submodules_len_modules.append(sub.find_element(By.TAG_NAME, "li"))
                # except:
                #     print(f"Couldn't find li: {s}")

        course_modules = driver.find_elements(By.CSS_SELECTOR, "span[data-tctest=section_name]")
        
      
        for e, li in enumerate (li_submodules_len_modules):
            submod_len = int(li.text.split(" ")[0])

            for i in range (submod_len):
                modules_data["modulo"].append(course_modules[e].text)
                # user_data[{course_modules[e].text}] 

                
                 
        # print(modules_data)           

        for s, sub_module in enumerate(sub_modules):
            # sub_module_title = sub_module.find_element()}

            file_type = sub_module.find_element(By.CSS_SELECTOR, "li[data-tctest=session_materials_type], li.no-files")
            modules_data["sub_modulo"].append(sub_module.text.replace(file_type.text, ""))
            modules_data["tipo de archivo"].append(file_type.text)


            if "VÍDEO" in file_type.text:
                file_type.click()
                preview_btn = driver.find_element(By.CSS_SELECTOR, "a[data-tctest=view_material_item]")
                preview_btn.click()
                video_duration = driver.find_element(By.CSS_SELECTOR, "span[id^=video-duration]")

                
                while not video_duration.text:
                    sleep(0.5)
                modules_data["duracion"].append(video_duration.text)
                print(f"Appending video duration: {video_duration.text}" )


                #print(video_duration.text)
                driver.back()
            else:
                modules_data["duracion"].append("No es video")
    
        #ir a pagina de informes 
        # obtener nombre de registrados
        # darle click a ver detalles para cada uno de los registrados
        # y obtener tiempo visto por cada modulo    
        
        # nos dirigimos a la pagina de informes en cada curso
        current_url = driver.current_url
        driver.get(current_url.replace("addcurriculum", "report/studentreport"))   

        # table rows
        table_rows = driver.find_elements(By.CSS_SELECTOR, "tr.ember-view")
    
        reg_data = {
            "nombre": [],
            "email": [],
            "fecha inscripcion": [],
            "last visit": []
        } 

        user_data = {
                "nombre": [],
                # "modulo": [],
                # "fecha inscripcion": [],
                # "last visit": []
            }
       
        for row in table_rows:
            tds = row.find_elements(By.TAG_NAME, "td")
            for k, td in enumerate(tds):
                if k == 1:
                    # Nombre
                    reg_data["nombre"].append(td.text)
                    print(f"Nombre: {td.text}")
                    user_data["nombre"].append(td.text)
                    
                if k == 2:
                    # email
                    reg_data["email"].append(td.text)
                    print(f"Correo: {td.text}")
                
                if k == 3:
                    # Fecha de inscripcion
                    reg_data["fecha inscripcion"].append(td.text)
                    print(f"Fecha de inscripcion: {td.text}")
         
                if k == 6:
                    #last
                    td.find_element(By.TAG_NAME, "a").click()
                    # handle click behaviour
                    last_login = driver.find_element(By.CSS_SELECTOR, "label[data-tctest=learner_details_last-login]")
                    reg_data["last visit"].append(last_login.text)
                    print(f"Last visit: {last_login.text}")
                    
                    

                    # capitulos
                    navigation_tabs = driver.find_elements(By.CSS_SELECTOR, "ul[role=Navigation]")
                    chapters = navigation_tabs[2].find_elements(By.CSS_SELECTOR, "*")

                    

                    total = 0

                    print(f"Cuantos chapters hay?: {len(chapters)}")
                    print(f"Driver url: {driver.current_url}")
                    
                    for element in chapters:
                        element.click()
                        driver.implicitly_wait(1)

                        time_spent = driver.find_elements(By.CSS_SELECTOR, "div[data-tctest='individual_lesson-material_result'] b")
                        sub_modules_titles = driver.find_elements(By.CSS_SELECTOR, "h6[data-tctest=individual_lesson_name]")

                        print("len(time_spent) =", len(time_spent))

                        id = 0
                        total_time_module = 0

                        for j, time in enumerate(time_spent):
                            id = id + 1   
                            print(f"Entered de loop {j}")
                            time = time_spent[j].text
                            time_seconds = parse_time(time)  
                            print(f"{time} -> {time_seconds}")
                               
                            
                            #   time = (int(time_spent[j].text.split.(" ")[5]))

                            # try:
                            #     time_seconds = (int(time_spent[j].text.split(":")[1])* 60) + (int(time_spent[j].text.split(":")[2]))
                            # except ValueError:
                            #     time_seconds = 0
                            
                            print(f"key: {sub_modules_titles[j].text}")
                            print(f"id: {id}")
                            try:
                                user_data[f"{sub_modules_titles[j].text} ({id})"].append(time_seconds)
                                total_time_module = total_time_module + time_seconds 
                            except KeyError:
                                user_data[f"{sub_modules_titles[j].text} ({id})"] = []
                                user_data[f"{sub_modules_titles[j].text} ({id})"].append(time_seconds)
                                total_time_module = total_time_module + time_seconds
                        total = total + total_time_module

                    driver.implicitly_wait(5)    
                    escape_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label=Close]")
                    escape_btn.click() 
                        
                        # try:
                        #     sub_modules = driver.find_elements(By.CSS_SELECTOR,"h6[data-tctest=individual_lesson_name]") 
                        #     print(f"lenght of submodules: {len(sub_modules)}")

                        #     for module in sub_modules:
                        #         module.click()
                        #         time_spent = driver.find_elements(By.CSS_SELECTOR,"div[data-tctest=individual_lesson-material_result] span")
                        #         total_time_module = 0
                        # except:
                        #     pass





                    # driver.implicitly_wait(0.2)  
                    # total = 0  
                    #user_data["Total"] = []

                    # id = 0
                    # for i in range(len(sections)):
                    #     sections = driver.find_elements(By.CSS_SELECTOR, ".reports-lesson-content.ember-view")
                    #     section = sections[i]
                    #     #user_data[sections_title[i].text] = []
                    #     section.click()

                    #     if not i == len(sections) - 1:
                    #         sleep(1)

                    #     sub_modules = section.find_elements(By.CSS_SELECTOR,"h6[data-tctest=individual_lesson_name]") 
                    #     print(f"lenght of submodules: {len(sub_modules)}")
                        
                    #     time_spent = section.find_elements(By.CSS_SELECTOR,"div[data-tctest=individual_lesson-material_result] span")
                    #     total_time_module = 0


                        # if (len(sub_modules) == 0):
                        #     user_data[sub.text].append(0)
                        


                        #user_data[sections_title[i].text].append(total_time_module)
                          
                    #user_data["Total"].append(total)       

                        #print([f"title: {sub_modules[i].text}, time: {time_spent[i].text}" for i in range(len(sub_modules))])
                   
                    

        
        wb = Workbook()

        #creates worksheets
        wb.create_sheet("Dataframe1")
        wb.create_sheet("Dataframe2")
        wb.create_sheet("Dataframe3")


        #access worksheets
        ws1 = wb["Dataframe1"]
        ws2 = wb["Dataframe2"]
        ws3 = wb["Dataframe3"]
        
        df1 = pd.DataFrame(reg_data, columns=reg_data.keys())

        print("ADDING TO DF2")
        lens = [f"{x}: {len(modules_data[x])}" for x in modules_data.keys()]
        print(lens)

        df2 = pd.DataFrame(modules_data, columns=modules_data.keys())
        print(user_data)

        df3 = pd.DataFrame(user_data, columns=user_data.keys())

        for r in dataframe_to_rows(df1, index=True, header=True):
            ws1.append(r)

        for r in dataframe_to_rows(df2, index=True, header=True):
            ws2.append(r)
        

        for r in dataframe_to_rows(df3, index=True, header=True):
            ws3.append(r)

        filename = f"Workbook for {course_title.text}.xlsx" 
        wb.save(filename)
        os.rename(filename, f"./xlsx/{filename}")
        
                    



        #nos regresamos al dashboard de cursos para ir a buscar el siguiente
        driver.get("https://ixlab-mx.trainercentral.com/adminpage#/portal/727165297/courses")

def findVideoLen():
    pass

if __name__ == "__main__":
    login()
    main()